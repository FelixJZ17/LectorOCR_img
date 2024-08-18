import os
import openpyxl

def escribir_excel(linea_a_escribir):

    # Separar el texto por " ; " para obtener una lista de líneas
    lineas = linea_a_escribir.split(' ; ')

    nombre_archivo = 'texto_en_excel.xlsx'

    # Si el archivo no existe, se crea un nuevo libro de Excel
    if os.path.exists(nombre_archivo):
        # Abre libro de Excel ya existente
        workbook = openpyxl.load_workbook(nombre_archivo)
    else:
        # Crear un nuevo libro de Excel
        workbook = openpyxl.Workbook()

    # Selecciona la hoja activa
    hoja = workbook.active
        
    fila = hoja.max_row + 1  # Encuentra la primera fila vacía

    # Escribir cada línea en una fila separada
    for idx, linea in enumerate(lineas, start=1):
        columna = idx  # Cada línea irá en una columna separada
        # Escribe el contenido en la celda correspondiente
        hoja.cell(row=fila, column=columna, value=linea)

    # Guardar el libro de Excel
    workbook.save(nombre_archivo)

    print(f"El texto ha sido escrito en {nombre_archivo}")